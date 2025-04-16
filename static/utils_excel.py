from openpyxl import load_workbook
import os

PADROES_CPFL = {
    "B1": {"disjuntor": "63A", "cabo": "16 mm²", "Caixa": "II", "Demanda": "18", "Fase": "Bifásico"},
    "B2": {"disjuntor": "70A", "cabo": "25 mm²", "Caixa": "II", "Demanda": "25", "Fase": "Bifásico"},
    "C1": {"disjuntor": "63A", "cabo": "16 mm²", "Caixa": "III", "Demanda": "23", "Fase": "Trifásico"},
    "C2": {"disjuntor": "80A", "cabo": "25 mm²", "Caixa": "III", "Demanda": "30", "Fase": "Trifásico"},
    "C3": {"disjuntor": "100A", "cabo": "35 mm²", "Caixa": "III", "Demanda": "38", "Fase": "Trifásico"},
    "C4": {"disjuntor": "125A", "cabo": "50 mm²", "Caixa": "H", "Demanda": "47", "Fase": "Trifásico"},
    "C5": {"disjuntor": "150A", "cabo": "70 mm²", "Caixa": "H", "Demanda": "57", "Fase": "Trifásico"},
    "C6": {"disjuntor": "200A", "cabo": "95 mm²", "Caixa": "H", "Demanda": "76", "Fase": "Trifásico"},
}

def preencher_anexo_f(
    modelo_path: str,
    saida_path: str,
    nome: str,
    cpf: str,
    telefone: str,
    email: str,
    endereco: str,
    cidade: str,
    latitude: str,
    longitude: str,
    data_operacao: str,
    qnt_mod: str,
    fab_mod: str,
    mod_mod:str,
    qnt_inv: str,
    fab_inv: str,
    mod_inv: str,
    pot_mod: str,
    pot_inv: str,
    pot_total: str,
    padrao: str,
    forma_conexao: str,
    uc:str,
    cep: str,
):
    # calculo pot kwp módulo
    pot_mod_watts = float(pot_mod)
    qnt_mod_int = int(qnt_mod)
    # Multiplica e converte para kWp
    pot_total_kwp = round((pot_mod_watts * qnt_mod_int) / 1000, 2)
    # calculo pot Kw inv
    pot_inv_kw = float (pot_inv)
    qnt_inv_int = int (qnt_inv)
    pot_inv = round((pot_inv_kw * qnt_inv_int) / 1000)

    area_mod = str(round(float(qnt_mod) * 2.5, 2))
    padrao_info = PADROES_CPFL.get(padrao, {})
    disjuntor = padrao_info.get("disjuntor", "")
    bitola = padrao_info.get("cabo", "")
    caixa = padrao_info.get("Caixa", "")
    demanda = padrao_info.get("Demanda", "")
    fase = padrao_info.get("Fase", "")

    wb = load_workbook(modelo_path)
    ws = wb.active

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                cell.value = (cell.value
                    .replace("#NOME", nome)
                    .replace("#UC", uc)
                    .replace("#CPF", cpf)
                    .replace("#TEL", telefone)
                    .replace("#CEL", email)
                    .replace("#ENDERECO", endereco)
                    .replace("#CIDADE", cidade)
                    .replace("#LAT", latitude)
                    .replace("#LONG", longitude)
                    .replace("#QNTMOD", qnt_mod)
                    .replace("#MODMOD", str(mod_mod))
                    .replace("#FABMOD", fab_mod)
                    .replace("#AREAMOD", area_mod)
                    .replace("#QNTINV", qnt_inv)
                    .replace("#FABINV", fab_inv)
                    .replace("#MODINV", mod_inv)
                    .replace("#POTMOD", str(pot_total_kwp))
                    .replace("#POTINV", str(pot_inv))
                    .replace("#DATA", data_operacao)
                    .replace("#KW", pot_total)
                    .replace("#PADRAO", padrao)
                    .replace("#CAIXA", caixa)
                    .replace("#DEMANDA", demanda)
                    .replace("#FASE", fase)
                    .replace("#TIPO", forma_conexao)         # aérea/subterrânea
                    .replace("#CABO", bitola)                # bitola correta
                    .replace("#DJ", disjuntor)               # disjuntor
                    .replace("#CEP", str(cep))
                )

    os.makedirs(os.path.dirname(saida_path), exist_ok=True)
    wb.save(saida_path)
